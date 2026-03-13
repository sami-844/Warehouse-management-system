"""
Product Management API endpoints
Create, Read, Update, Delete products
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from typing import List, Optional

from app.core.database import get_db
from app.models.product import Product, ProductCategory, DeletedItemsLog
from app.models.user import User
from datetime import datetime, timezone
import json
from app.schemas.product import (
    Product as ProductSchema,
    ProductCreate,
    ProductUpdate,
    ProductCategory as CategorySchema,
    ProductCategoryCreate,
    ProductCategoryUpdate
)
from app.api.auth import get_current_user

router = APIRouter()


# ==================== PRODUCT CATEGORIES ====================

@router.get("/categories", response_model=List[CategorySchema])
def list_categories(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, le=1000),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get all product categories
    
    - **skip**: Number of records to skip (for pagination)
    - **limit**: Maximum number of records to return
    """
    categories = db.query(ProductCategory).offset(skip).limit(limit).all()
    return categories


@router.post("/categories", response_model=CategorySchema, status_code=201)
def create_category(
    category: ProductCategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Create a new product category
    
    Requires authentication
    """
    # Check if category name already exists
    existing = db.query(ProductCategory).filter(ProductCategory.name == category.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Category name already exists")
    
    db_category = ProductCategory(**category.dict())
    db.add(db_category)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Category name already exists")
    db.refresh(db_category)
    return db_category


@router.get("/categories/{category_id}", response_model=CategorySchema)
def get_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get a specific category by ID
    """
    category = db.query(ProductCategory).filter(ProductCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return category


@router.put("/categories/{category_id}", response_model=CategorySchema)
def update_category(
    category_id: int,
    category_update: ProductCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a product category."""
    db_cat = db.query(ProductCategory).filter(ProductCategory.id == category_id).first()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Category not found")
    update_data = category_update.dict(exclude_unset=True)
    if "name" in update_data and update_data["name"] != db_cat.name:
        existing = db.query(ProductCategory).filter(ProductCategory.name == update_data["name"]).first()
        if existing:
            raise HTTPException(status_code=400, detail="Category name already exists")
    for field, value in update_data.items():
        setattr(db_cat, field, value)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Category name already exists")
    db.refresh(db_cat)
    return db_cat


@router.delete("/categories/{category_id}", status_code=204)
def delete_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Soft-delete a product category. Blocks if products are assigned."""
    db_cat = db.query(ProductCategory).filter(ProductCategory.id == category_id).first()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Category not found")
    product_count = db.query(Product).filter(Product.category_id == category_id, Product.is_active == True).count()
    if product_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {product_count} active products assigned")
    db_cat.is_active = False
    db.commit()
    return None


# ==================== PRODUCTS ====================

@router.get("/products", response_model=List[ProductSchema])
def list_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, le=1000),
    category_id: Optional[int] = None,
    search: Optional[str] = None,
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get all products with optional filters
    
    - **skip**: Number of records to skip (for pagination)
    - **limit**: Maximum number of records to return
    - **category_id**: Filter by category
    - **search**: Search in product name or SKU
    - **is_active**: Filter by active status
    """
    query = db.query(Product).filter(Product.is_deleted != True)

    # Apply filters
    if category_id is not None:
        query = query.filter(Product.category_id == category_id)
    
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            (Product.name.ilike(search_filter)) | 
            (Product.sku.ilike(search_filter)) |
            (Product.barcode.ilike(search_filter))
        )
    
    if is_active is not None:
        query = query.filter(Product.is_active == is_active)
    
    products = query.offset(skip).limit(limit).all()
    return products


@router.post("/products", response_model=ProductSchema, status_code=201)
def create_product(
    product: ProductCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Create a new product
    
    Requires authentication
    """
    # Check if SKU already exists
    existing = db.query(Product).filter(Product.sku == product.sku).first()
    if existing:
        raise HTTPException(status_code=400, detail="SKU already exists")
    
    # Check if barcode already exists (if provided)
    if product.barcode:
        existing_barcode = db.query(Product).filter(Product.barcode == product.barcode).first()
        if existing_barcode:
            raise HTTPException(status_code=400, detail="Barcode already exists")
    
    # Create product — coerce nulls to defaults for non-nullable fields
    product_data = product.dict()
    # Extract opening stock fields before creating product
    opening_qty = int(product_data.pop('opening_qty', 0) or 0)
    opening_cost = float(product_data.pop('opening_cost', 0) or 0)
    if product_data.get('tax_rate') is None:
        product_data['tax_rate'] = 0.00
    if product_data.get('reorder_level') is None:
        product_data['reorder_level'] = 10
    if product_data.get('minimum_stock') is None:
        product_data['minimum_stock'] = 5
    # Remove brand_id if present (not a DB column yet in some environments)
    product_data.pop('brand_id', None)
    db_product = Product(**product_data, created_by=current_user.id)
    db.add(db_product)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="SKU or barcode already exists")
    db.refresh(db_product)

    # Create opening stock transaction if qty > 0
    if opening_qty > 0:
        try:
            from app.models.inventory import InventoryTransaction, StockLevel
            unit_cost = opening_cost if opening_cost > 0 else float(db_product.standard_cost or 0)
            txn = InventoryTransaction(
                product_id=db_product.id,
                warehouse_id=1,
                transaction_type='RECEIPT',
                quantity=opening_qty,
                unit_cost=unit_cost,
                total_cost=round(unit_cost * opening_qty, 3),
                reference_type='opening_stock',
                reference_number=f"OPEN-{db_product.sku}",
                notes='Opening stock on product creation',
                created_by=current_user.id,
            )
            db.add(txn)
            # Update or create stock level
            sl = db.query(StockLevel).filter(
                StockLevel.product_id == db_product.id,
                StockLevel.warehouse_id == 1
            ).first()
            if sl:
                sl.quantity_on_hand = (sl.quantity_on_hand or 0) + opening_qty
                sl.quantity_available = (sl.quantity_available or 0) + opening_qty
                sl.total_value = float(sl.quantity_on_hand or 0) * unit_cost
                sl.average_cost = unit_cost
            else:
                sl = StockLevel(
                    product_id=db_product.id,
                    warehouse_id=1,
                    quantity_on_hand=opening_qty,
                    quantity_reserved=0,
                    quantity_available=opening_qty,
                    total_value=round(unit_cost * opening_qty, 3),
                    average_cost=unit_cost,
                )
                db.add(sl)
            # Update product standard_cost if opening_cost was provided
            if opening_cost > 0:
                db_product.standard_cost = opening_cost
            db.commit()
            db.refresh(db_product)
        except Exception as e:
            # Don't fail the whole product creation if stock txn fails
            db.rollback()
            print(f"Warning: Opening stock transaction failed: {e}")

    return db_product


@router.get("/products/avg-costs")
def product_avg_costs(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Average cost per product from purchase order items."""
    from sqlalchemy import text
    from app.core.database import engine
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT poi.product_id,
                       ROUND(COALESCE(SUM(poi.quantity * poi.unit_price), 0) /
                             NULLIF(COALESCE(SUM(poi.quantity), 0), 0), 3) as avg_cost,
                       COALESCE(SUM(poi.quantity), 0) as total_qty
                FROM purchase_order_items poi
                GROUP BY poi.product_id
            """))
            keys = rows.keys()
            data = {r[0]: {"avg_cost": float(r[1] or 0), "total_qty": int(r[2] or 0)}
                    for r in rows.fetchall()}
        return {"costs": data}
    except Exception:
        return {"costs": {}}


@router.get("/products/{product_id}/stock")
def get_product_stock(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get current stock level for a product."""
    from app.models.inventory import StockLevel
    sl = db.query(StockLevel).filter(
        StockLevel.product_id == product_id,
        StockLevel.warehouse_id == 1
    ).first()
    if sl:
        return {
            "quantity_on_hand": float(sl.quantity_on_hand or 0),
            "quantity_reserved": float(sl.quantity_reserved or 0),
            "quantity_available": float(sl.quantity_available or 0),
            "average_cost": float(sl.average_cost or 0),
        }
    return {"quantity_on_hand": 0, "quantity_reserved": 0, "quantity_available": 0, "average_cost": 0}


@router.get("/products/deleted")
def list_deleted_products(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all deleted items from the audit log."""
    items = db.query(DeletedItemsLog).filter(
        DeletedItemsLog.is_restored == False
    ).order_by(DeletedItemsLog.deleted_at.desc()).all()
    return [
        {
            "id": i.id, "item_type": i.item_type, "item_id": i.item_id,
            "item_name": i.item_name, "item_sku": i.item_sku,
            "item_data": i.item_data,
            "deleted_by_id": i.deleted_by_id, "deleted_by_name": i.deleted_by_name,
            "deleted_at": i.deleted_at.isoformat() if i.deleted_at else None,
            "deleted_reason": i.deleted_reason,
            "is_restored": i.is_restored,
        }
        for i in items
    ]


@router.get("/products/{product_id}", response_model=ProductSchema)
def get_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get a specific product by ID
    """
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.get("/products/sku/{sku}", response_model=ProductSchema)
def get_product_by_sku(
    sku: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get a specific product by SKU
    
    Useful for barcode scanning operations
    """
    product = db.query(Product).filter(Product.sku == sku).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.get("/products/barcode/{barcode}", response_model=ProductSchema)
def get_product_by_barcode(
    barcode: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get a specific product by barcode
    
    Useful for barcode scanning operations
    """
    product = db.query(Product).filter(Product.barcode == barcode).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.put("/products/{product_id}", response_model=ProductSchema)
def update_product(
    product_id: int,
    product_update: ProductUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Update an existing product
    
    Only provided fields will be updated
    """
    # Get existing product
    db_product = db.query(Product).filter(Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Update only provided fields
    update_data = product_update.dict(exclude_unset=True)
    
    # Check for SKU conflict — exclude current product
    if "sku" in update_data and update_data["sku"]:
        existing = db.query(Product).filter(
            Product.sku == update_data["sku"],
            Product.id != product_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="SKU already used by another product")

    # Check for barcode conflict — exclude current product
    if "barcode" in update_data and update_data["barcode"]:
        existing = db.query(Product).filter(
            Product.barcode == update_data["barcode"],
            Product.id != product_id,
            Product.barcode != ''
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Barcode already used by another product")

    # Strip unknown fields the model doesn't have
    update_data.pop('brand_id', None)

    # Coerce nullable-false fields to safe defaults
    if 'tax_rate' in update_data:
        update_data['tax_rate'] = float(update_data['tax_rate'] or 0)
    if 'reorder_level' in update_data:
        update_data['reorder_level'] = int(update_data['reorder_level'] or 0)
    if 'minimum_stock' in update_data:
        update_data['minimum_stock'] = int(update_data['minimum_stock'] or 0)
    if 'maximum_stock' in update_data:
        update_data['maximum_stock'] = int(update_data['maximum_stock'] or 0)

    # Update fields
    for field, value in update_data.items():
        if hasattr(db_product, field):
            setattr(db_product, field, value)

    db_product.updated_by = current_user.id
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Update failed: {str(e)[:200]}")
    db.refresh(db_product)
    return db_product


@router.delete("/products/{product_id}")
def delete_product(
    product_id: int,
    reason: str = Query("", description="Reason for deletion"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Soft delete a product with audit trail."""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Soft delete the product
    product.is_deleted = True
    product.is_active = False
    product.deleted_at = datetime.now(timezone.utc)
    product.deleted_by = current_user.id
    product.deleted_reason = reason or None
    product.updated_by = current_user.id

    # Log in deleted_items_log
    log_entry = DeletedItemsLog(
        item_type="product",
        item_id=product.id,
        item_name=product.name,
        item_sku=product.sku,
        item_data=json.dumps({
            "id": product.id, "name": product.name, "sku": product.sku,
            "barcode": product.barcode,
            "selling_price": str(product.selling_price or 0),
            "standard_cost": str(product.standard_cost or 0),
            "unit_of_measure": product.unit_of_measure,
            "category_id": product.category_id,
            "description": product.description,
            "reorder_level": product.reorder_level,
            "tax_rate": str(product.tax_rate or 0),
            "is_perishable": product.is_perishable,
        }),
        deleted_by_id=current_user.id,
        deleted_by_name=current_user.username,
        deleted_at=datetime.now(timezone.utc),
        deleted_reason=reason or None,
    )
    db.add(log_entry)
    db.commit()

    return {"message": f"Product '{product.name}' deleted successfully"}


@router.post("/products/{product_id}/restore")
def restore_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Restore a soft-deleted product."""
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if not product.is_deleted:
        raise HTTPException(status_code=400, detail="Product is not deleted")

    # Restore the product
    product.is_deleted = False
    product.is_active = True
    product.deleted_at = None
    product.deleted_by = None
    product.deleted_reason = None
    product.updated_by = current_user.id

    # Update the log entry
    log = db.query(DeletedItemsLog).filter(
        DeletedItemsLog.item_id == product_id,
        DeletedItemsLog.item_type == "product",
        DeletedItemsLog.is_restored == False
    ).order_by(DeletedItemsLog.deleted_at.desc()).first()
    if log:
        log.is_restored = True
        log.restored_at = datetime.now(timezone.utc)
        log.restored_by_id = current_user.id
        log.restored_by_name = current_user.username

    db.commit()
    return {"message": f"Product '{product.name}' restored successfully"}


@router.post("/products/import", status_code=201)
def import_products(
    rows: List[dict],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Bulk import products from CSV/Excel rows. Only 'name' is required. Auto-generates SKU if missing."""
    import uuid
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        try:
            name = str(row.get('name', '') or '').strip()
            if not name:
                skipped += 1
                continue
            sku = str(row.get('sku', '') or '').strip()
            if not sku:
                sku = f"SKU-{str(uuid.uuid4())[:8].upper()}"
            if db.query(Product).filter(Product.sku == sku).first():
                skipped += 1
                continue
            barcode = str(row.get('barcode', '') or '').strip() or None
            if barcode and db.query(Product).filter(Product.barcode == barcode).first():
                barcode = None
            p = Product(
                sku=sku, name=name,
                unit_of_measure=str(row.get('unit', 'PCS') or 'PCS').strip(),
                selling_price=float(row.get('selling_price', 0) or 0),
                standard_cost=float(row.get('standard_cost', 0) or 0),
                barcode=barcode,
                description=str(row.get('description', '') or '').strip() or None,
                is_active=True, created_by=current_user.id
            )
            db.add(p)
            created += 1
        except Exception as e:
            errors.append(f"Row {i+1}: {str(e)}")
            skipped += 1
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Import failed — database constraint error")
    return {"created": created, "skipped": skipped, "errors": errors[:10]}


@router.post("/products/import-file", status_code=201)
async def import_products_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Import products from CSV or Excel file.
    Required columns: sku, name
    Optional: selling_price, standard_cost, unit_of_measure, category,
              opening_quantity, tax_rate, reorder_level, minimum_stock, barcode
    """
    import uuid, csv, io, traceback as tb
    from app.models.inventory import InventoryTransaction

    def safe_float(val, default=0):
        try:
            return float(val) if val else default
        except (ValueError, TypeError):
            return default

    try:
        content = await file.read()
        fname = (file.filename or '').lower()

        rows = []
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            if len(data) < 2:
                return {"imported": 0, "skipped": 0, "errors": ["File has no data rows"]}
            headers = [str(h or '').strip().lower().replace(' ', '_') for h in data[0]]
            for vals in data[1:]:
                row = {}
                for j, h in enumerate(headers):
                    row[h] = str(vals[j]) if j < len(vals) and vals[j] is not None else ''
                rows.append(row)
            wb.close()
        elif fname.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append({k.strip().lower().replace(' ', '_'): (v or '').strip() for k, v in r.items()})
        else:
            raise HTTPException(400, 'Only CSV and Excel (.xlsx) files are supported')

        imported, skipped, errors = 0, 0, []
        for i, row in enumerate(rows, start=2):
            try:
                row = {str(k).lower().strip(): str(v).strip() if v is not None else ''
                       for k, v in row.items()}

                sku = row.get('sku', '').strip()
                name = row.get('name', '').strip()
                if not name:
                    skipped += 1
                    continue
                if not sku:
                    sku = f"SKU-{str(uuid.uuid4())[:8].upper()}"
                if db.query(Product).filter(Product.sku == sku).first():
                    skipped += 1
                    continue

                # Auto-create category if provided
                category_id = None
                cat_name = row.get('category', '').strip()
                if cat_name:
                    cat = db.query(ProductCategory).filter(
                        ProductCategory.name.ilike(cat_name)
                    ).first()
                    if not cat:
                        cat = ProductCategory(name=cat_name, is_active=True)
                        db.add(cat)
                        db.flush()
                    category_id = cat.id

                barcode = row.get('barcode', '').strip() or None
                if barcode and db.query(Product).filter(Product.barcode == barcode).first():
                    barcode = None

                selling_price = safe_float(row.get('selling_price') or row.get('sell_price'))
                standard_cost = safe_float(row.get('standard_cost') or row.get('cost_price'))
                opening_qty = safe_float(row.get('opening_quantity') or row.get('opening_qty'))

                p = Product(
                    sku=sku, name=name,
                    category_id=category_id,
                    unit_of_measure=row.get('unit_of_measure') or row.get('unit') or 'pcs',
                    selling_price=selling_price,
                    standard_cost=standard_cost,
                    tax_rate=safe_float(row.get('tax_rate'), 5),
                    reorder_level=safe_float(row.get('reorder_level')),
                    minimum_stock=safe_float(row.get('minimum_stock')),
                    barcode=barcode,
                    description=row.get('description', '').strip() or None,
                    is_active=True, is_perishable=False,
                    created_by=current_user.id
                )
                db.add(p)
                db.flush()

                # Create opening stock transaction if qty > 0
                if opening_qty > 0:
                    try:
                        txn = InventoryTransaction(
                            product_id=p.id,
                            warehouse_id=1,
                            transaction_type='RECEIPT',
                            quantity=opening_qty,
                            unit_cost=standard_cost,
                            total_cost=round(standard_cost * opening_qty, 3),
                            reference_type='opening_stock',
                            reference_number=f"OPEN-{sku}",
                            notes=f'Opening stock imported for {name}',
                            created_by=current_user.id,
                        )
                        db.add(txn)
                    except Exception:
                        pass  # don't fail import if stock txn fails

                imported += 1
            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")
                skipped += 1

        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            raise HTTPException(status_code=400, detail="Import failed — database constraint error")
        return {
            "imported": imported, "skipped": skipped,
            "errors": errors[:10],
            "message": f"Import complete: {imported} products added, {skipped} skipped"
        }
    except HTTPException:
        raise
    except Exception as e:
        tb.print_exc()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")