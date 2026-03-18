"""Suppliers CRUD API"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
from app.core.database import get_db
from app.api.auth import get_current_user
from app.models.user import User
from app.models.business_partner import Supplier
from app.models.purchase import PurchaseOrder, PurchaseInvoice

router = APIRouter()

class SupplierCreate(BaseModel):
    code: str
    name: str
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    mobile: Optional[str] = None
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    postal_code: Optional[str] = None
    payment_terms_days: int = 30
    credit_limit: Optional[float] = None
    tax_id: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    notes: Optional[str] = None
    opening_balance: Optional[float] = None
    vendor_type: Optional[str] = "supplier"

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    mobile: Optional[str] = None
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    payment_terms_days: Optional[int] = None
    credit_limit: Optional[float] = None
    tax_id: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None
    vendor_type: Optional[str] = None

@router.get("")
async def list_suppliers(active_only: bool = False, vendor_type: Optional[str] = None, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    query = db.query(Supplier)
    if active_only:
        query = query.filter(Supplier.is_active == True)
    if vendor_type:
        query = query.filter(Supplier.vendor_type == vendor_type)
    suppliers = query.order_by(Supplier.name).all()
    result = []
    for s in suppliers:
        po_count = db.query(func.count(PurchaseOrder.id)).filter(PurchaseOrder.supplier_id == s.id).scalar() or 0
        total_ordered = db.query(func.sum(PurchaseOrder.total_amount)).filter(PurchaseOrder.supplier_id == s.id).scalar() or 0
        outstanding = db.query(func.sum(PurchaseInvoice.total_amount - PurchaseInvoice.amount_paid)).filter(
            PurchaseInvoice.supplier_id == s.id, PurchaseInvoice.status != 'paid'
        ).scalar() or 0
        result.append({
            "id": s.id, "code": s.code, "name": s.name,
            "contact_person": s.contact_person, "email": s.email,
            "phone": s.phone, "mobile": s.mobile,
            "address_line1": s.address_line1, "city": s.city, "country": s.country,
            "payment_terms_days": s.payment_terms_days,
            "credit_limit": float(s.credit_limit) if s.credit_limit else None,
            "tax_id": s.tax_id, "bank_name": s.bank_name, "bank_account": s.bank_account,
            "notes": s.notes, "is_active": s.is_active,
            "vendor_type": getattr(s, 'vendor_type', 'supplier') or 'supplier',
            "total_orders": po_count,
            "total_ordered_value": round(float(total_ordered), 3),
            "outstanding_balance": round(float(outstanding), 3),
        })
    return result

@router.get("/{supplier_id}")
async def get_supplier(supplier_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")
    orders = db.query(PurchaseOrder).filter(PurchaseOrder.supplier_id == supplier_id).order_by(PurchaseOrder.order_date.desc()).limit(20).all()
    invoices = db.query(PurchaseInvoice).filter(PurchaseInvoice.supplier_id == supplier_id).order_by(PurchaseInvoice.invoice_date.desc()).limit(20).all()
    return {
        "id": s.id, "code": s.code, "name": s.name,
        "contact_person": s.contact_person, "email": s.email,
        "phone": s.phone, "mobile": s.mobile,
        "address_line1": s.address_line1, "address_line2": s.address_line2,
        "city": s.city, "country": s.country, "postal_code": s.postal_code,
        "payment_terms_days": s.payment_terms_days,
        "credit_limit": float(s.credit_limit) if s.credit_limit else None,
        "tax_id": s.tax_id, "bank_name": s.bank_name, "bank_account": s.bank_account,
        "notes": s.notes, "is_active": s.is_active,
        "orders": [{"id": o.id, "po_number": o.po_number, "date": str(o.order_date), "status": o.status, "total": float(o.total_amount) if o.total_amount else 0} for o in orders],
        "invoices": [{"id": i.id, "number": i.invoice_number, "date": str(i.invoice_date), "total": float(i.total_amount), "paid": float(i.amount_paid), "status": i.status} for i in invoices],
    }

@router.post("", status_code=201)
async def create_supplier(data: SupplierCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if db.query(Supplier).filter(Supplier.code == data.code).first():
        raise HTTPException(status_code=400, detail="Supplier code already exists")
    opening_bal = float(data.opening_balance or 0)
    fields = data.dict(exclude={'opening_balance'})
    if opening_bal > 0:
        existing_notes = fields.get('notes') or ''
        fields['notes'] = f"[Opening Balance: {opening_bal:.3f} OMR]\n{existing_notes}".strip()
    s = Supplier(**fields, created_by=current_user.id)
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": s.id, "code": s.code, "name": s.name, "opening_balance": opening_bal, "message": "Supplier created"}

@router.put("/{supplier_id}")
async def update_supplier(supplier_id: int, data: SupplierUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")
    for field, value in data.dict(exclude_unset=True).items():
        setattr(s, field, value)
    s.updated_by = current_user.id
    db.commit()
    return {"id": s.id, "name": s.name, "message": "Supplier updated"}


@router.post("/import", status_code=201)
async def import_suppliers(rows: List[dict], db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Bulk import suppliers from CSV/Excel rows. No required fields — auto-generates code and name if missing."""
    import uuid
    from sqlalchemy.exc import IntegrityError
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        try:
            name = str(row.get('name', '') or '').strip()
            code = str(row.get('code', '') or '').strip()
            if not name and not code:
                skipped += 1
                continue
            if not name:
                name = f"Supplier-{code}"
            if not code:
                code = f"SUP-{str(uuid.uuid4())[:8].upper()}"
            if db.query(Supplier).filter(Supplier.code == code).first():
                skipped += 1
                continue
            s = Supplier(
                code=code, name=name,
                contact_person=str(row.get('contact_person', '') or '').strip() or None,
                phone=str(row.get('phone', '') or '').strip() or None,
                email=str(row.get('email', '') or '').strip() or None,
                city=str(row.get('city', '') or '').strip() or None,
                payment_terms_days=int(row.get('payment_terms_days', 30) or 30),
                is_active=True, created_by=current_user.id
            )
            db.add(s)
            created += 1
        except Exception as e:
            errors.append(f"Row {i+1}: {str(e)}")
            skipped += 1
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Import failed — database constraint error")
    return {"created": created, "skipped": skipped, "errors": errors[:10]}


@router.post("/import-file", status_code=201)
async def import_suppliers_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Import suppliers from an uploaded CSV or Excel file."""
    import uuid, csv, io
    from sqlalchemy.exc import IntegrityError
    content = await file.read()
    fname = (file.filename or '').lower()

    rows = []
    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if len(data) < 2:
            return {"created": 0, "skipped": 0, "errors": ["File has no data rows"]}
        headers = [str(h or '').strip().lower().replace(' ', '_') for h in data[0]]
        for vals in data[1:]:
            row = {}
            for j, h in enumerate(headers):
                row[h] = str(vals[j]) if j < len(vals) and vals[j] is not None else ''
            rows.append(row)
        wb.close()
    else:
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append({k.strip().lower().replace(' ', '_'): (v or '').strip() for k, v in r.items()})

    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        try:
            name = str(row.get('name', '') or '').strip()
            code = str(row.get('code', '') or '').strip()
            if not name and not code:
                skipped += 1
                continue
            if not name:
                name = f"Supplier-{code}"
            if not code:
                code = f"SUP-{str(uuid.uuid4())[:8].upper()}"
            if db.query(Supplier).filter(Supplier.code == code).first():
                skipped += 1
                continue
            s = Supplier(
                code=code, name=name,
                contact_person=str(row.get('contact_person', '') or '').strip() or None,
                phone=str(row.get('phone', '') or '').strip() or None,
                email=str(row.get('email', '') or '').strip() or None,
                city=str(row.get('city', '') or '').strip() or None,
                payment_terms_days=int(row.get('payment_terms_days', 30) or 30),
                is_active=True, created_by=current_user.id
            )
            db.add(s)
            created += 1
        except Exception as e:
            errors.append(f"Row {i+1}: {str(e)}")
            skipped += 1
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Import failed — database constraint error")
    return {"created": created, "skipped": skipped, "errors": errors[:10]}


# ===== SUPPLIER PRICE LISTS =====
@router.get("/{supplier_id}/price-list")
async def get_supplier_price_list(supplier_id: int, db: Session = Depends(get_db),
                                  current_user: User = Depends(get_current_user)):
    """Get all product prices for a specific supplier."""
    from sqlalchemy import text as sql_text
    from app.core.database import engine
    with engine.connect() as conn:
        rows = conn.execute(sql_text("""
            SELECT spl.id, spl.product_id, p.name as product_name, p.sku,
                   spl.unit_price, spl.min_order_qty, spl.lead_time_days,
                   spl.notes, spl.is_active, spl.last_updated
            FROM supplier_price_list spl
            JOIN products p ON spl.product_id = p.id
            WHERE spl.supplier_id = :sid
            ORDER BY p.name
        """), {"sid": supplier_id}).fetchall()
    return [{
        "id": r[0], "product_id": r[1], "product_name": r[2], "sku": r[3],
        "unit_price": round(float(r[4] or 0), 3), "min_order_qty": float(r[5] or 1),
        "lead_time_days": r[6], "notes": r[7] or "", "is_active": r[8],
        "last_updated": str(r[9]) if r[9] else None,
    } for r in rows]


@router.post("/{supplier_id}/price-list")
async def upsert_supplier_price(supplier_id: int, data: dict, db: Session = Depends(get_db),
                                current_user: User = Depends(get_current_user)):
    """Add or update a product price for a supplier."""
    from sqlalchemy import text as sql_text
    from app.core.database import engine
    product_id = data.get("product_id")
    unit_price = float(data.get("unit_price", 0))
    min_qty = float(data.get("min_order_qty", 1))
    lead_days = int(data.get("lead_time_days", 7))
    notes = data.get("notes", "")

    with engine.connect() as conn:
        conn.execute(sql_text("""
            INSERT INTO supplier_price_list (supplier_id, product_id, unit_price, min_order_qty, lead_time_days, notes, updated_by, last_updated)
            VALUES (:sid, :pid, :price, :min_qty, :lead, :notes, :uid, NOW())
            ON CONFLICT (supplier_id, product_id)
            DO UPDATE SET unit_price = :price, min_order_qty = :min_qty, lead_time_days = :lead,
                          notes = :notes, updated_by = :uid, last_updated = NOW()
        """), {"sid": supplier_id, "pid": product_id, "price": unit_price,
               "min_qty": min_qty, "lead": lead_days, "notes": notes, "uid": current_user.id})
        conn.commit()
    return {"message": "Price saved", "supplier_id": supplier_id, "product_id": product_id, "unit_price": unit_price}


@router.post("/{supplier_id}/price-list/bulk")
async def bulk_upsert_prices(supplier_id: int, data: dict, db: Session = Depends(get_db),
                             current_user: User = Depends(get_current_user)):
    """Bulk add/update prices for a supplier."""
    from sqlalchemy import text as sql_text
    from app.core.database import engine
    items = data.get("items", [])
    saved = 0
    with engine.connect() as conn:
        for item in items:
            try:
                conn.execute(sql_text("""
                    INSERT INTO supplier_price_list (supplier_id, product_id, unit_price, min_order_qty, lead_time_days, notes, updated_by, last_updated)
                    VALUES (:sid, :pid, :price, :min_qty, :lead, :notes, :uid, NOW())
                    ON CONFLICT (supplier_id, product_id)
                    DO UPDATE SET unit_price = :price, min_order_qty = :min_qty, lead_time_days = :lead,
                                  notes = :notes, updated_by = :uid, last_updated = NOW()
                """), {
                    "sid": supplier_id, "pid": item.get("product_id"),
                    "price": float(item.get("unit_price", 0)),
                    "min_qty": float(item.get("min_order_qty", 1)),
                    "lead": int(item.get("lead_time_days", 7)),
                    "notes": item.get("notes", ""), "uid": current_user.id
                })
                saved += 1
            except Exception:
                pass
        conn.commit()
    return {"message": f"Saved {saved} prices for supplier", "saved": saved}


@router.get("/price-lookup")
async def price_lookup(supplier_id: int, product_id: int, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    """Quick lookup: get the price for a product from a specific supplier."""
    from sqlalchemy import text as sql_text
    from app.core.database import engine
    with engine.connect() as conn:
        row = conn.execute(sql_text("""
            SELECT unit_price, min_order_qty, lead_time_days
            FROM supplier_price_list
            WHERE supplier_id = :sid AND product_id = :pid AND is_active = true
        """), {"sid": supplier_id, "pid": product_id}).fetchone()
    if not row:
        return {"found": False}
    return {"found": True, "unit_price": round(float(row[0]), 3), "min_order_qty": float(row[1]), "lead_time_days": row[2]}
